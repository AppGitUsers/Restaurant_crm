from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
import datetime

from apps.accounts.permissions import IsAdmin
from .models import Transaction, Expense, DailyReport
from .serializers import TransactionSerializer, ExpenseSerializer, DailyReportSerializer


class TransactionViewSet(viewsets.ModelViewSet):
    queryset           = Transaction.objects.all()
    serializer_class   = TransactionSerializer
    permission_classes = [IsAdmin]
    filterset_fields   = ['tx_type', 'category', 'tx_date']
    search_fields      = ['description', 'reference']
    ordering_fields    = ['tx_date', 'amount', 'created_at']

    @action(detail=True, methods=['post'], url_path='secure_delete')
    def secure_delete(self, request, pk=None):
        password = request.data.get('password', '')
        if not password:
            return Response({'error': 'Password is required.'}, status=400)
        if not request.user.check_password(password):
            return Response({'error': 'Incorrect password.'}, status=400)
        self.get_object().delete()
        return Response(status=204)


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset           = Expense.objects.all()
    serializer_class   = ExpenseSerializer
    permission_classes = [IsAdmin]
    filterset_fields   = ['category', 'expense_date']
    search_fields      = ['title', 'notes']
    ordering_fields    = ['expense_date', 'amount']

    def perform_create(self, serializer):
        expense = serializer.save()
        Transaction.objects.create(
            tx_type     = 'EXPENSE',
            amount      = expense.amount,
            category    = expense.category,
            description = expense.title,
            reference   = f"expense:{expense.id}",
        )


class DailyReportViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = DailyReport.objects.all()
    serializer_class   = DailyReportSerializer
    permission_classes = [IsAdmin]
    filterset_fields   = ['report_date']
    ordering_fields    = ['report_date', 'total_sales']


class FinanceSummaryView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        today       = timezone.now().date()
        month_start = today.replace(day=1)

        total_income  = Transaction.objects.filter(tx_type='INCOME').aggregate(t=Sum('amount'))['t'] or 0
        total_expense = Transaction.objects.filter(tx_type='EXPENSE').aggregate(t=Sum('amount'))['t'] or 0

        month_income  = Transaction.objects.filter(tx_type='INCOME',  tx_date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0
        month_expense = Transaction.objects.filter(tx_type='EXPENSE', tx_date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0

        today_income  = Transaction.objects.filter(tx_type='INCOME',  tx_date=today).aggregate(t=Sum('amount'))['t'] or 0
        today_expense = Transaction.objects.filter(tx_type='EXPENSE', tx_date=today).aggregate(t=Sum('amount'))['t'] or 0

        # GST breakdown for the current month
        month_tax_income = Transaction.objects.filter(
            tx_type='INCOME', tx_date__gte=month_start
        ).aggregate(t=Sum('tax_amount'))['t'] or 0
        month_base_income = float(month_income) - float(month_tax_income)

        today_tax_income = Transaction.objects.filter(
            tx_type='INCOME', tx_date=today
        ).aggregate(t=Sum('tax_amount'))['t'] or 0

        # Last 30 days daily breakdown
        thirty_days_ago = today - datetime.timedelta(days=30)
        daily_data = (Transaction.objects
                      .filter(tx_date__gte=thirty_days_ago)
                      .annotate(date=TruncDate('tx_date'))
                      .values('date', 'tx_type')
                      .annotate(total=Sum('amount'))
                      .order_by('date'))

        # Monthly breakdown (last 12 months)
        twelve_months_ago = today - datetime.timedelta(days=365)
        monthly_data = (Transaction.objects
                        .filter(tx_date__gte=twelve_months_ago)
                        .annotate(month=TruncMonth('tx_date'))
                        .values('month', 'tx_type')
                        .annotate(total=Sum('amount'))
                        .order_by('month'))

        # Category breakdown for expenses
        expense_by_category = (Transaction.objects
                                .filter(tx_type='EXPENSE')
                                .values('category')
                                .annotate(total=Sum('amount'))
                                .order_by('-total'))

        return Response({
            'summary': {
                'total_income':      float(total_income),
                'total_expense':     float(total_expense),
                'net_profit':        float(total_income) - float(total_expense),
                'month_income':      float(month_income),
                'month_expense':     float(month_expense),
                'month_profit':      float(month_income) - float(month_expense),
                'month_tax_income':  float(month_tax_income),
                'month_base_income': float(month_base_income),
                'today_income':      float(today_income),
                'today_expense':     float(today_expense),
                'today_tax_income':  float(today_tax_income),
            },
            'daily_chart':       list(daily_data),
            'monthly_chart':     list(monthly_data),
            'expense_breakdown': list(expense_by_category),
        })


class MonthlyReportView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        import calendar
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from apps.billing.models import Order

        year  = int(request.query_params.get('year',  timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        orders = Order.objects.filter(
            status='PAID',
            created_at__year=year,
            created_at__month=month,
        ).order_by('created_at')

        expense_txns = Transaction.objects.filter(
            tx_type='EXPENSE',
            tx_date__year=year,
            tx_date__month=month,
        ).order_by('tx_date')

        wb = openpyxl.Workbook()

        # ── Styles ──────────────────────────────────────────────────
        h_font  = Font(bold=True, color='FFFFFF', size=11)
        h_green = PatternFill(fill_type='solid', fgColor='1D9E75')
        h_gold  = PatternFill(fill_type='solid', fgColor='BA7517')
        t_font  = Font(bold=True, size=11)
        t_green = PatternFill(fill_type='solid', fgColor='D9F2EC')
        t_gold  = PatternFill(fill_type='solid', fgColor='FEF3CD')
        center  = Alignment(horizontal='center')
        right   = Alignment(horizontal='right')

        def style_header(ws, headers, fill):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = h_font
                c.fill = fill
                c.alignment = center

        def style_total(ws, total_row, ncols, fill):
            for col in range(1, ncols + 1):
                c = ws.cell(row=total_row, column=col)
                c.font = t_font
                c.fill = fill

        def set_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # ── Income Sheet ────────────────────────────────────────────
        ws_i = wb.active
        ws_i.title = 'Income'

        inc_headers = [
            'Bill ID', 'Customer Name', 'Customer Phone',
            'Bill Amount (₹)', 'Base Amount (₹)', 'GST %', 'GST Amount (₹)',
            'Payment Method', 'Date',
        ]
        style_header(ws_i, inc_headers, h_green)

        tot_bill = tot_base = tot_tax = 0.0
        for r, order in enumerate(orders, 2):
            tax  = float(order.tax_amount)
            base = float(order.total_amount) - tax
            ws_i.cell(r, 1, order.order_number)
            ws_i.cell(r, 2, order.customer_name or 'Walk-in')
            ws_i.cell(r, 3, order.customer_phone or '—')
            ws_i.cell(r, 4, float(order.total_amount)).alignment = right
            ws_i.cell(r, 5, round(base, 2)).alignment = right
            ws_i.cell(r, 6, float(order.tax_percent)).alignment = right
            ws_i.cell(r, 7, round(tax, 2)).alignment = right
            ws_i.cell(r, 8, order.get_payment_method_display())
            ws_i.cell(r, 9, str(order.created_at.date()))
            tot_bill += float(order.total_amount)
            tot_base += base
            tot_tax  += tax

        tr = orders.count() + 2
        style_total(ws_i, tr, len(inc_headers), t_green)
        ws_i.cell(tr, 1, 'TOTAL')
        ws_i.cell(tr, 4, round(tot_bill, 2)).alignment = right
        ws_i.cell(tr, 5, round(tot_base, 2)).alignment = right
        ws_i.cell(tr, 7, round(tot_tax,  2)).alignment = right
        set_widths(ws_i, [16, 22, 16, 16, 16, 8, 16, 16, 12])

        # ── Expense Sheet ────────────────────────────────────────────
        ws_e = wb.create_sheet('Expenses')

        exp_headers = ['Category', 'Description', 'Amount (₹)', 'Date']
        style_header(ws_e, exp_headers, h_gold)

        tot_exp = 0.0
        exp_list = list(expense_txns)
        for r, tx in enumerate(exp_list, 2):
            ws_e.cell(r, 1, tx.category)
            ws_e.cell(r, 2, tx.description)
            ws_e.cell(r, 3, float(tx.amount)).alignment = right
            ws_e.cell(r, 4, str(tx.tx_date))
            tot_exp += float(tx.amount)

        er = len(exp_list) + 2
        style_total(ws_e, er, len(exp_headers), t_gold)
        ws_e.cell(er, 1, 'TOTAL')
        ws_e.cell(er, 3, round(tot_exp, 2)).alignment = right
        set_widths(ws_e, [20, 40, 16, 12])

        # ── Stream response ─────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        month_name = calendar.month_name[month]
        filename   = f"finance_{month_name}_{year}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp


class MonthlyDataView(APIView):
    """JSON version of the monthly report — powers the in-page income/expense tables."""
    permission_classes = [IsAdmin]

    def get(self, request):
        from apps.billing.models import Order

        year  = int(request.query_params.get('year',  timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        orders = Order.objects.filter(
            status='PAID',
            created_at__year=year,
            created_at__month=month,
        ).order_by('-created_at')

        # Map order reference -> Transaction.id for deletion
        refs   = [f"order:{o.order_number}" for o in orders]
        tx_map = {
            tx.reference: tx.id
            for tx in Transaction.objects.filter(reference__in=refs, tx_type='INCOME')
        }

        income_rows = []
        for order in orders:
            ref  = f"order:{order.order_number}"
            tax  = float(order.tax_amount)
            base = float(order.total_amount) - tax
            income_rows.append({
                'tx_id':          tx_map.get(ref),
                'order_number':   order.order_number,
                'customer_name':  order.customer_name or 'Walk-in',
                'customer_phone': order.customer_phone or '—',
                'bill_amount':    float(order.total_amount),
                'base_amount':    round(base, 2),
                'gst_percent':    float(order.tax_percent),
                'gst_amount':     round(tax, 2),
                'payment_method': order.get_payment_method_display(),
                'date':           str(order.created_at.date()),
            })

        expense_txns = Transaction.objects.filter(
            tx_type='EXPENSE',
            tx_date__year=year,
            tx_date__month=month,
        ).order_by('-tx_date')

        expense_rows = [
            {
                'tx_id':       tx.id,
                'category':    tx.category,
                'description': tx.description,
                'amount':      float(tx.amount),
                'date':        str(tx.tx_date),
            }
            for tx in expense_txns
        ]

        return Response({'income': income_rows, 'expenses': expense_rows})
